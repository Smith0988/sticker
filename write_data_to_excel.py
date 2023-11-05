import os
import shutil
import sys

import openpyxl
from openpyxl.styles import Alignment


def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)


temp_excel_form = resource_path("temp_data\\temp_form.xlsx")


def write_to_excel(data):
    excel_file = "temp_form.xlsx"
    if not os.path.exists(excel_file):
        source_excel = resource_path("temp_data\\temp_form.xlsx")
        shutil.copy(source_excel, excel_file)
    workbook = openpyxl.load_workbook(excel_file, data_only=True)
    sheet = workbook["Sheet1"]
    next_row = sheet.max_row + 1

    col_1_data = data[0]
    col_2_data = data[1]
    col_3_data = data[2]
    col_4_data = data[3]
    sheet.cell(row=next_row, column=1, value=col_1_data).alignment = Alignment(horizontal='center')
    sheet.cell(row=next_row, column=2, value=col_2_data).alignment = Alignment(horizontal='center')
    sheet.cell(row=next_row, column=3, value=col_3_data).alignment = Alignment(horizontal='center')
    sheet.cell(row=next_row, column=4, value=col_4_data).alignment = Alignment(horizontal='center')

    # Lưu tệp Excel
    workbook.save(excel_file)


if __name__ == "__main__":
    # Sử dụng hàm để ghi dữ liệu vào tệp Excel
    data_to_write = ["value1", "value2", "value3", "value4"]
    write_to_excel(data_to_write)
